import openpyxl
from sys import stderr
import os.path
import pathlib


rfc_path = pathlib.Path(r"C:\Users\nick\Downloads\Desktop")
print(rfc_path) 



reimbursable_pay_items = ["1-4.","1-6.","2-3.","4-1.b.","14-3.","15-3.","15-3.a.","15-3.b.","15-7.",
                          "15-8.","15-9.","15-10.","22-1.","22-2.","22-3.","22-4.","22-5.","22-6.",
                          "22-7.","22-8.","22-9.","22-10.","22-11.","22-12.","22-13.","22-14.","22-15.",
                          "22-16.","22-17.","22-18.","22-19.","22-20.","22-21.","22-22.","22-23.","22-24.",
                          "22-25.","22-26.","22-27.","22-28.","22-29.","22-30.",]
def get_pay_items():
    pay_items = ['1.', '1-1.', '1-2.', '1-2.a.', '1-3.', '1-4.', '1-5.',
             '1-5.a.', '1-6.', '1-7.', '2.', '2-1.', '2-2.', '2-3.',
             '2-4.', '3.', '3-1.', '3-2.', '3-3.', '3-4.', '3-5.',
             '3-6.', '3-7.a.', '3-8.a.', '3-9.a.', '3-10.a.', '3-11.',
             '3-12.', '3-13.a.', '3-14.a.', '3-15.', '3-16.', '3-17.',
             '3-18.', '3-19.', '3-20.', '4.', '4-1.a.', '4-1.b.', '5.',
             '5-1.a.1.', '5-1.a.2.', '5-1.a.3.', '5-1.b.', '5-2.', '5-3.a.',
             '5-5.a.', '5-6.', '5-7.', '5-8.', '5-9.', '5-10.', '5-11.',
             '5-12.', '5-13.', '5-14.', '5-15.', '5-16.', '5-17.', '5-18.',
             '5-19.', '5-20.', '5-21.', '5-22.', '5-23.', '6.', '6-1.',
             '6-2.a.', '6-2.b.', '6-3.a.', '6-3.b.', '6-4.', '6-5.', '6-6.',
             '6-7.', '6-8.', '6-9.a.', '6-9.b.', '6-9.c.', '6-9.d.', '6-10.',
             '6-11.', '6-12.', '6-13.', '6-14.', '6-15.', '7.', '7-1.', '7-2.',
             '7-3.', '7-4.', '7-5.', '7-6.', '7-7.', '8.', '8-1.', '8-2.', '8-3.',
             '8-4.', '8-5.', '8-6.', '8-7.', '8-8.', '8-9.', '8-10.', '8-11.',
             '8-12.', '8-13.', '8-14.', '9.', '9.A.', '9-1.', '9-2.', '9-3.',
             '9-4.', '9-5.', '9-6.', '9-7.', '9-8.', '9-8.a.', '9-9.', '9-10.',
             '9-11.', '9-12.', '9-13.', '9-14.', '9-15.', '9-16.', '9-17.', '9-18.',
             '9-19.', '9-20.', '9-21.', '9-22.', '9-23.', '9-24.', '9.B.', '9-25.',
             '9-26.', '9-27.', '9-28.', '9-29.', '9-30.', '9-31.', '9-31.a.', '9-32.',
             '9-33.', '9-34.', '9-35.', '9-36.', '9-37.', '9-38.', '9-39.', '9-40.',
             '9-41.', '9-41.a.', '9-42.', '9-43.', '9-44.', '9-44.a.', '9-45.', '9-46.',
             '9-47.', '9-48.', '9-49.', '9-50.', '9-51.', '9-52.', '9-53.', '9-54.', '9-55.', 
             '9-56.', '9-57.', '9-58.', '9-59.', '9-60.', '9-61.', '9-62.', '9-63.', '9-64.', 
             '9-65.', '9-66.', '9-67.', '9-78.', '9-79.', '9-80.', '9.C.', '9-68.', '9-69.', 
             '9-70.', '9-71.', '9.D.', '9-72.', '9-73.', '9.E.', '9-75.', '9-76.', '9-77.', 
             '10.', '10-1.a.', '10-1.b.', '10-1.c.', '10-1.d.', '10-2.a.', '10-2.b.', '10-2.c.', 
             '10-2.d.', '10-7.', '10-8.', '10-9.', '10-10.', '10-11.a.', '10-11.b.', '10-12.a.', 
             '10-12.b.', '10-13.', '10-14.', '10-15.', '10-15.a.', '10-15.b.', '10-16.', 
             '10-17.', '10-18.', '10-19.', '10-20.', '10-21.', '10-22.', '10-23.', '11.', 
             '11-1.', '11-2.', '11-3.', '11-4.', '12.', '12-1.', '12-2.', '12-3.', 
             '12-4.', '12-5.', '12-6.', '12-7.', '12-8.', '12-9.', '12-10.', '12-11.', 
             '12-12.', '12-13.', '12-14.', '12-15.', '12-16.', '12-17.', '12-18.', '13.', 
             '13-1.', '13-2.', '13-3.', '13-4.', '13-5.', '13-6.', '13-7.', '14.', '14-1.a.',
             '14-2.a.', '14-3.', '14-4.', '14-5.', '15.', '15.A.', '15-1.a.', '15-1.b.',
             '15-1.c.', '15-1.d.', '15-2.a.', '15-3.', '15-3.a.', '15-3.b.', '15.B.', 
             '15-4.a.', '15-4.b.', '15-4.c.', '15-4.d.', '15-5.', '15-7.', '15-8.',
             '15-9.', '15-10.', '16.', '16.A.', '16-1.', '16-2.', '16-3.', '16-4.',
             '16-5.', '16-6.', '16-7.', '16-8.', '16-9.', '16-10.', '16-11.', '16-12.', 
             '16-13.', '16-14.', '16.B.', '16-17.', '16-18.', '16-19.', '16-20.', '16-21.', 
             '16-22.', '16-23.', '16-24.', '16-25.', '16-26.', '16-27.', '16-28.', '17.', 
             '17-1.', '17-2.', '17-3.', '17-4.', '17-5.', '18.', '18-1.', '18-2.', '18-3.', 
             '18-4.', '18-5.', '18-6.', '18-7.', '18-8.', '18-9.', '18-10.', '18-11.', 
             '18-12.', '18-13.', '18-14.', '18-15.', '18-16.', '18-17.', '18-18.', '18-19.', 
             '18-20.', '18-21.', '18-22.', '18-23.', '18-24.', '18-25.', '18-26.', '18-27.',
             '18-28.', '18-29.', '18-30.', '18-31.', '18-32.', '18-33.', '18-34.', '18-35.', 
             '18-36.', '18-37.', '18-38.', '18-39.', '18-40.', '18-41.', '18-42.', '18-43.', 
             '18-44.', '18-45.', '18-46.', '18-47.', '18.48.', '18-49.', '18-50.', '18-51.', 
             '18-52.', '18-53.', '18-54.', '18-55.', '18-56.', '18-57.', '18-58.', '19.', '19-1.', 
             '19-3.', '19-4.', '19-5.', '19-6.', '19-7.', '19-8.', '19-9.', '19-10.', 
             '19-11.', '19-12.', '19-13.', '19-14.', '19-15.', '19-16.', '19-17.', '19-18.',
             '19-19.', '19-20.', '19-21.', '19-22.', '19-23.', '19-24.', '19-25.', '19-26.',
             '19-27.', '20.', '20-1.', '20-2.', '20-3.', '20-4.', '20-5.', '20-6.', '20-7.',
             '20-8.', '20-9.', '20-10.', '20-11.', '21.', '21-1.', '21-2.', '21-3.', '21-4.',
             '21-5.', '21-6.a.', '21-6.b.', '21-6.c.', '21-6.d.', '21-7.a.', '21-7.b.',
             '21-8.', '21-9.', '21-10.', '21-11.', '21-12.', '21-13.', '21-15.', '21-16.',
             '21-17.', '21-18.', '21-19.', '21-20.', '21-21.', '21-22.', '21-23.',
             '21-24.', '21-25.', '21-26.', '21-27.', '21-28.', '21-29.', '21-30.',
             '21-31.', '21-32.', '21-33.', '21-34.', '21-35.', '21-36.', '22.', 
             '22-1.', '22-2.', '22-3.', '22-4.', '22-5.', '22-6.', '22-7.', '22-8.',
             '22-9.', '22-10.', '22-11.', '22-12.', '22-13.', '22-14.', '22-15.', '22-16.',
             '22-17.', '22-18.', '22-19.', '22-20.', '22-21.', '22-22.', '22-23.',
             '22-24.', '22-25.', '22-26.', '22-27.', '22-28.', '22-29.', '22-30.', '23.', '23-1.']
    return pay_items
def get_tasks():
    tasks = ['NEGOTIATED ITEM PRICE', 'TOTAL QUANTITIES','TASK 1', 
         'TASK 2','TASK 3', 'TASK 4', 'TASK 5', 'TASK 6', 
         'TASK 7', 'TASK 8', 'TASK 9', 'TASK 10']
    return tasks
def my_input(prompt=None):
    '''this function replaces input() and prints prompt to stderr'''
    if prompt:
        stderr.write(str(prompt))
    return input()
    
def sort_by_pi(e):
    '''this is a helper function which is used to sort the list of rfc
    entries into the same order as the list of payitems as they appear in the SPI'''
    return get_pay_items().index(e[1])

def load_rfc(rfc_worksheet):
    '''this takes a worksheet from RFC file and creates a list of tuples, 
    each tuple includes tsk, pi, uom, pi_price, new_val, total_price'''
    lol = []
    for row in rfc_worksheet.iter_rows(min_row=22, max_row=111, min_col=0,max_col=9, values_only=True):
        if row[0] == None: continue
        tsk, pi, desc,_,_, uom, pi_price, new_val, total_price = row
        lol.append(('TASK '+str(tsk),pi, uom, new_val, total_price))
    lol.sort()
    lol.sort(key=sort_by_pi)
    return lol   

def get_spi():
    spi_wb = my_input('enter rfc to update: ')
    spi_wb = pathlib.Path(spi_wb) if spi_wb != '' else  r"C:\Users\nick\Downloads\Desktop\PART2-AttachmentB-SPI-138506266-SA - Copy.xlsm" #example: C:/Users/e317181/OneDrive - Miami-Dade County/Desktop/AttachmentB(Revision1)-SPI-138505875-SA.xlsm
    print(spi_wb)
    return openpyxl.load_workbook(spi_wb, keep_vba=True, keep_links=True)

def get_rfc():
    rfc_wb = my_input('enter rfc: ')
    rfc_wb = pathlib.Path(rfc_wb) if (rfc_wb != '') else pathlib.Path(r"C:\Users\nick\Downloads\Desktop\5632 RFC1 VERSION 4 - bad.xlsm")
    # example: C:/Users/e317181/OneDrive - Miami-Dade County/Desktop/rfc.xlsm
    print(rfc_wb)
    return openpyxl.load_workbook(rfc_wb, read_only=True, data_only=True, keep_links=True)

def get_spi_row(pi):
    min_row=11
    return min_row + get_pay_items().index(pi)

def get_spi_column(task):
    min_col=7
    return min_col + get_tasks().index(task)

def get_old_spi_cell(pi, task):
    row = get_spi_row(pi)
    column = get_spi_column(task)
    cell = SOW_Units.cell(row=row, column=column) 
    cell.value = cell.value if cell.value != None else 0
    return cell

#def add_to_spi (oldspicell, rfcvalue):
#    new_val = oldspicell.value + rfcvalue
#    SOW_Units.cell(row=row, column=column, value=new_val)

  

rfc = get_rfc()
spi = get_spi()
                            #spi['Invoice'].defined_names.clear()
rfc_worksheet = rfc['RFC']
SOW_Units = spi['SOW Units']


rfc_items = load_rfc(rfc_worksheet)
handling_fee = False
for rfc_item in rfc_items:
    task,pi,units = rfc_item[0],rfc_item[1],rfc_item[3]
    oldval = get_old_spi_cell(pi, task).value
    get_old_spi_cell(pi, task).value += units
    print('{} unit(s) added to {} in {}, {} The new value \
          is {}'.format(units, 
                        oldval, 
                        task,
                        pi, 
                        get_old_spi_cell(pi, task).value))
    
    SOW_Units.row_dimensions[get_spi_row(pi)].hidden = False
    
if handling_fee: SOW_Units.row_dimensions[20].hidden = False   # unhide the 6% handling fee row. 
spi.save(rfc_path / 'newgeneratedspi.xlsm')
print("new spi saved to the same folder as the rfc")
spi = openpyxl.load_workbook(rfc_path / 'newgeneratedspi.xlsm',read_only=True, data_only=True)
input('press enter to quit')

#if I add an item make sure its row is visible
#get oldspivalue and add new quantity to it.


